from __future__ import print_function
from openpyxl import load_workbook
import httplib2
import os

from apiclient import discovery
from oauth2client import client
from oauth2client import tools
from oauth2client.file import Storage

import datetime
from dateutil import parser

#=====================================================================================================================================================================


#	LOCAL SETUP


class mevent:
	def __init__(self, date,event_type,lid):
		self.date = date
		self.event_type = event_type
		self.lid = lid
	def show(self):
		output = ''
		output = '{}{}'.format(output,self.event_type)
		output = '{} of {}'.format(output,self.lid)		
		output = '{} at {}'.format(output,self.date.strftime('%m/%d/%y'))
		return output
	def equals(self,me):
		if (not me.date.month == self.date.month) or (not me.date.day == self.date.day) or (not me.date.year == self.date.year) :
			return False
		if not me.event_type == self.event_type:
			return False
		if not me.lid == self.lid:
			return False
		return True

wb = load_workbook('MouseInformation_AllLines_copy.xlsx')

def getCol(s,t):
	for cell in s[1]:
		if cell.value is not None:
			if cell.value.startswith(t):
				return cell.column
	return 0

def eventsfcol(des,lid,dates):
	mevents = []
	for cell in dates:
			if cell.value is not None:
				if type(cell.value) is type(datetime.datetime.now()):
					id = lid[cell.row-1]
					e = mevent(cell.value,des,id.value)
					mevents.append(e)
	return mevents				



local = []

for ws in wb:
	if getCol(ws,'Dissection Date') is not 0 and getCol(ws, 'Litter ID') is not 0:
		disevents = eventsfcol('Dissect',ws[getCol(ws,'Litter ID')], ws[getCol(ws,'Dissection Date')])
		for event in disevents:
			local.append(event)

	if getCol(ws,'Wean') is not 0 and getCol(ws, 'Litter ID') is not 0:
		winevents = eventsfcol('Wean',ws[getCol(ws,'Litter ID')], ws[getCol(ws,'Wean')])
		for event in winevents:
			local.append(event)

	if getCol(ws,'Tattoo') is not 0 and getCol(ws, 'Litter ID') is not 0:
		tatevents = eventsfcol('Tattoo',ws[getCol(ws,'Litter ID')], ws[getCol(ws,'Tattoo')])
		for event in tatevents:
			local.append(event)


#=====================================================================================================================================================================



#	API SETUP



try:
    import argparse
    flags = argparse.ArgumentParser(parents=[tools.argparser]).parse_args()
except ImportError:
    flags = None

# If modifying these scopes, delete your previously saved credentials
# at ~/.credentials/calendar-python-quickstart.json
SCOPES = 'https://www.googleapis.com/auth/calendar'
CLIENT_SECRET_FILE = 'client_secret.json'
APPLICATION_NAME = 'Google Calendar API Python Quickstart'


def get_credentials():
    """Gets valid user credentials from storage.

    If nothing has been stored, or if the stored credentials are invalid,
    the OAuth2 flow is completed to obtain the new credentials.

    Returns:
        Credentials, the obtained credential.
    """
    home_dir = os.path.expanduser('~')
    credential_dir = os.path.join(home_dir, '.credentials')
    if not os.path.exists(credential_dir):
        os.makedirs(credential_dir)
    credential_path = os.path.join(credential_dir,
                                   'calendar-python-quickstart.json')

    store = Storage(credential_path)
    credentials = store.get()
    if not credentials or credentials.invalid:
        flow = client.flow_from_clientsecrets(CLIENT_SECRET_FILE, SCOPES)
        flow.user_agent = APPLICATION_NAME
        if flags:
            credentials = tools.run_flow(flow, store, flags)
        else: # Needed only for compatibility with Python 2.6
            credentials = tools.run(flow, store)
        print('Storing credentials to ' + credential_path)
    return credentials



credentials = get_credentials()
http = credentials.authorize(httplib2.Http())
service = discovery.build('calendar', 'v3', http=http)


gevents = {}

page_token = None
while True:
	events = service.events().list(calendarId='6khpu5g1ut54ihpbfe1em9jbsc@group.calendar.google.com', pageToken=page_token).execute()
	for event in events['items']:
		ed = parser.parse(event['start']['dateTime']).date()
		gevent = mevent(ed, event['summary'].split(' ')[0], event['summary'].split(' ')[1])	
		gevents[event['id']] = gevent
  	page_token = events.get('nextPageToken')
  	if not page_token:
  	  break
print('')


#uploading missing events:

for event in local:
	eventExists = False

	for gevent in gevents:
		if gevents[gevent].equals(event):
			eventExists = True
	
	if not eventExists:
		des = '{} {}'.format(event.event_type,event.lid)
		sdt = parser.parse('{} 9:00 AM'.format(event.date.isoformat())).isoformat()
		edt = parser.parse('{} 11:00 AM'.format(event.date.isoformat())).isoformat()
		ge = {
			'summary': des,
			'start': {
				'dateTime': sdt,
				'timeZone': 'America/New_York'
			},
			'end': {
				'dateTime': edt,
			 	'timeZone': 'America/New_York'
			},
			'reminders': {
				'useDefault': True
			}	
		}
		e = service.events().insert(calendarId='6khpu5g1ut54ihpbfe1em9jbsc@group.calendar.google.com', body=ge).execute()


#delete non-existing events:

for event in gevents:
	eventExists = False

	for levent in local:
		if levent.equals(gevents[event]):
			eventExists = True
	
	if not eventExists:
		service.events().delete(calendarId='6khpu5g1ut54ihpbfe1em9jbsc@group.calendar.google.com', eventId=event).execute()
	



"""
now = datetime.datetime.utcnow().isoformat() + 'Z' # 'Z' indicates UTC time
print('Getting the upcoming 10 events')
eventsResult = service.events().list(
calendarId='primary', timeMin=now, maxResults=10, singleEvents=True,
orderBy='startTime').execute()
events = eventsResult.get('items', [])

if not events:
	print('No upcoming events found.')
for event in events:
	start = event['start'].get('dateTime', event['start'].get('date'))
	print(start, event['summary'])
"""



"""
event = {
  'summary': 'Google I/O 2015',
  'location': '800 Howard St., San Francisco, CA 94103',
  'description': 'A chance to hear more about Google\'s developer products.',
  'start': {
    'dateTime': '2015-05-28T09:00:00-07:00',
    'timeZone': 'America/Los_Angeles',
  },
  'end': {
    'dateTime': '2015-05-28T17:00:00-07:00',
    'timeZone': 'America/Los_Angeles',
  },
  'recurrence': [
    'RRULE:FREQ=DAILY;COUNT=2'
  ],
  'attendees': [
    {'email': 'lpage@example.com'},
    {'email': 'sbrin@example.com'},
  ],
  'reminders': {
    'useDefault': False,
    'overrides': [
      {'method': 'email', 'minutes': 24 * 60},
      {'method': 'popup', 'minutes': 10},
    ],
  },
}

event = service.events().insert(calendarId='primary', body=event).execute()
#print 'Event created: %s' % (event.get('htmlLink'))

"""



"""
try:
	import argparseu
	flags = argparse.ArgumentParser(parents=[tools.argparser]).parse_args()
except ImportError:
	flags = None

SCOPES = 'https://www.googleapis.com/auth/calendar'
CLIENT_SECRET = 'client_secret.json'

store = file.Storage('storage.json')
credz = store.get()

if not credz or credz.invalid:
	flow = client.flow_from_clientsecrets(CLIENT_SECRET, SCOPES)
	credz = tools.run_flow(flow, store, flags) \ 
		if flags else tools.run(flow,store)

API_KEY = 'AIzaSyCTswttl1PbzHtbD5oy3v7Cj1iILgAiI8k'

CAL = build('calendar', 'v3', http=credz.authorize(Http()))

GMT_OFF = '-04:00'

EVENT = {
	'summary': 'Dissect mouse',
	'start': {'dateTime': '2018-08-2BT19:00:00%s' % GMT_OFF},
	'end': {'dateTime': '2018-08-2BT22:00:00%s' % GMT_OFF}
}

"""
